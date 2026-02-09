import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

# 1. Read Excel file
df = pd.read_excel("expenses.xlsx")

# 2. Analyze data: sum by category
# Dictionary to store total amount per category
category_totals = {}

# Go through each row in the table
for index, row in df.iterrows():
    category = row["category"]
    amount = row["amount"]

    if category in category_totals:
        category_totals[category] += amount
    else:
        category_totals[category] = amount

summary = pd.DataFrame(
    list(category_totals.items()),
    columns=["category", "amount"]
)


# 3. Save analysis to a new Excel file
output_file = "finance_analysis.xlsx"
summary.to_excel(output_file, index=False, sheet_name="Summary")

# 4. Open Excel with openpyxl to add a chart
wb = load_workbook(output_file)
ws = wb["Summary"]

# 5. Create a bar chart
chart = BarChart()
chart.title = "Витрати за категоріями"
chart.x_axis.title = "Категорія"
chart.y_axis.title = "Сума"

data = Reference(
    ws,
    min_col=2,
    min_row=1,
    max_row=ws.max_row
)

categories = Reference(
    ws,
    min_col=1,
    min_row=2,
    max_row=ws.max_row
)

chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)

# 6. Place chart in Excel
ws.add_chart(chart, "D2")

# 7. Save file
wb.save(output_file)

print("Готово! Створено файл finance_analysis.xlsx з діаграмою.")
